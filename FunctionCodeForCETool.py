import pandas as pd
import math 
import re
import openpyxl
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess
import os
import traceback

#----------------------------------- Excell File Loading Functions------------------------------------------------#de
def load_excel_range(filename, sheet_name, end_row, end_col):
    df = pd.read_excel(filename, sheet_name=sheet_name, engine='openpyxl', header=None)
    df = df.iloc[:end_row, :end_col]
    df.columns = df.iloc[0]
    df = df.iloc[1:]
    return df
#___Cleaning____#
def cleanInchFeet(value):
    val = value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.replace("'", "").replace('"', '').strip()
        if val == '':
            return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

#-----------------------Matching Function-----------------------------------------------------------#
##### To be Used In finding Standard Types or Pipes names?
def Find_Match(MatchItem,Match_list):
    CleanItem = str(MatchItem).lower().replace('"','').replace("'", "").replace("-", " ")
    for InList in Match_list:
        ListClean = str(InList).lower().replace('"','').replace("'", "").replace("-", " ")
        if ListClean in CleanItem:
            return InList
    return None
    
#----------------------------------------------------------Structure Functions------------------------------------------------------------------------------#

def StructureCalcs(row,PI_df,CB_df,INL_df,MH_df,log=print):
        
    cb_standard_types = CB_df['Type'].dropna().tolist()
    inl_SPU_types =INL_df['Type'].dropna().tolist()
    mh_standard_types = MH_df['Type'].dropna().tolist()

    name = str(row['Name']).strip()
    description = str(row['Description']).strip().lower()
    row_type = str(row['Type']).strip()

    if "null structure" in row_type.lower():
        return ([name, row_type, 0, 0, 0, 0, 0, 0])
    ################ FIND LOWEST PIPE CONECTING TO STRUCTURE TO DETERMINE MAX_PIPE DEPTH########################
    target_name = name.strip().upper()
    normalized_starts = PI_df['Start Structure'].astype(str).str.strip().str.upper()    
    normalized_ends = PI_df['End Structure'].astype(str).str.strip().str.upper() 
    matching_start = PI_df[normalized_starts == target_name]
    matching_end = PI_df[normalized_ends == target_name]
      
    if not matching_start.empty or matching_end.empty:
            
            lowest_invert = min(matching_start['Start Invert Elevation'].astype(float).min(), matching_end['End Invert Elevation'].astype(float).min())
            rim_elevation = float(row['Insertion Rim Elevation'])
            max_pipe_depth = round(rim_elevation - lowest_invert, 2)
    else:
        max_pipe_depth = 0
    ############################### MAX PIPE DEPTH FINISH #######################################################
    ###################START LOOKING FOR MATCHES TO GENERATE CALCULATIONS#########################
    matched_type = None
    match_df = None
    keyword = None

    # Determine if CB or MH  and skip if EX in name
    if 'ex' in name.lower():
        log(f"skipping existing structre {name}")
        return([name, row_type, 0, 0, 0, 0, 0, 0])
    
    if 'cb' in name.lower():
        matched_type = Find_Match(description, cb_standard_types)
        
        if matched_type:
            match_df = CB_df[CB_df['Type'] == matched_type]
            keyword = 'cb'
    elif 'inl' in name.lower():
        matched_type = Find_Match(description, inl_SPU_types)
        
        if matched_type:
            match_df = INL_df[INL_df['Type'] == matched_type]
            keyword = 'cb'    
    elif 'mh' in name.lower():
        matched_type = Find_Match(description, mh_standard_types)
        
        if matched_type:
            match_df = MH_df[MH_df['Type'] == matched_type]
            keyword = 'mh'

    if match_df is not None and not match_df.empty:
        area_of_structure = 0
        structure_depth = 0
        extra_area = 0
        excavation = 0
        shoring = 0

        # CB Calculations
        if keyword == 'cb':
            
            shape = match_df['Shape'].values[0].strip().capitalize()
            length = float(match_df['Length/Diameter'].values[0])
            width = float(match_df['Width'].values[0])
            wall_thickness = float(match_df['Wall Thickness'].values[0])
            base_thickness = float(match_df['Base Thickness'].values[0])
            sump_thickness = float(match_df['Sump Thickness'].values[0])
            bedding_thickness = float(match_df['Bedding Thickness'].values[0])

            structure_depth = max_pipe_depth + (base_thickness + sump_thickness + bedding_thickness) / 12
        #########################SQUARE CATCH BASIN CALCS#####################################################################################
            if shape == 'Square':
                area_of_structure = (length + 2 * wall_thickness) * (width + 2 * wall_thickness) / 144
                extra_area = ((length + 24) * (width + 24)) / 144 - area_of_structure
                #if structure_depth > 4:
                #    shoring_length = length + width + wall_thickness * 4
                #    shoring = (shoring_length / 12) * structure_depth
        ###########################LARGER CATCH BASINS CALCS#############################################################################
            elif shape == 'Circular':
                radius = (length + 2 * wall_thickness) / 2
                area_of_structure = math.pi * (radius ** 2) / 144
                extra_radius = (length + 2 * wall_thickness + 24) / 2
                extra_area = math.pi * (extra_radius ** 2) / 144 - area_of_structure
                #if structure_depth > 4:
                #    shoring_length = math.pi * (length + wall_thickness * 2) / 2
                #   shoring = (shoring_length / 12) * structure_depth

            excavation = (area_of_structure + extra_area) * structure_depth / 27
            StructureBeddingArea = (area_of_structure + extra_area) * bedding_thickness / 27

        # MH Calculations
        elif keyword == 'mh':
            
            diameter = float(match_df['Diameter'].values[0])
            wall_thickness = float(match_df['Wall Thickness'].values[0])
            base_thickness = float(match_df['Base Thickness'].values[0])
            bedding_thickness = float(match_df['Bedding Thickness'].values[0])

            radius = (diameter + 2 * wall_thickness) / 2
            area_of_structure = math.pi * (radius ** 2) / 144
            structure_depth = max_pipe_depth + (base_thickness + bedding_thickness) / 12
            extra_radius = (diameter + 2 * wall_thickness + 24) / 2
            extra_area = math.pi * (extra_radius ** 2) / 144 - area_of_structure

            excavation = (area_of_structure + extra_area) * structure_depth / 27
            StructureBeddingArea = (area_of_structure + extra_area) * bedding_thickness / 27
            if structure_depth > 4:
                shoring_length = math.pi * (diameter + wall_thickness * 2) / 2
                shoring = (shoring_length / 12) * structure_depth
        else:
             log(" No CB/MH/INL match — skipping processing for this structure.")            
        return ([ name, row['Description'], round(max_pipe_depth,2), round(area_of_structure,2),round(structure_depth,2), round(extra_area,2), round(excavation,2),round(StructureBeddingArea,2) ])
    else: 
        log(f"Structure Error {name} Values Defaulted to Zero")
        return([name, row_type, 0, 0, 0, 0, 0, 0])
    


#-------------------------------------------------------------------Pipe Calculation Function --------------------------------------------------------------------------
def PipeCalc(row,SI_df,material_keywords,log =print):
    name = row['Name']
    desc = row['Description']
    dia = row['Inner Diameter']  # in inches
    length = row['2D Length']  # in feet
    slope = row['Slope']
    start_struct = row['Start Structure']
    end_struct = row['End Structure']
    start_inv = row['Start Invert Elevation']
    end_inv = row['End Invert Elevation']
    start_cov = row['Start Cover']
    if start_cov == 0:
            matching_structure_row = SI_df[SI_df['Name'].str.strip().str.upper() == str(start_struct).strip().upper()] 
            if not matching_structure_row.empty:
                InsertE =     matching_structure_row.iloc[0].get('Insertion Rim Elevation')
                InsertE = cleanInchFeet(InsertE)
                start_cov = InsertE - start_inv      
    end_cov = row['End Cover']
    if end_cov == 0:
            matching_structure_row = SI_df[SI_df['Name'].str.strip().str.upper() == str(end_struct).strip().upper()]
            if not matching_structure_row.empty:
                InsertE = matching_structure_row.iloc[0].get('Insertion Rim Elevation')
                InsertE = cleanInchFeet(InsertE)
                end_cov = InsertE - end_inv

    # Skip row if critical fields are missing
    if None in (dia, length, start_cov, end_cov):
        log("Missing required field(s) for Pipe Calculations")

    #Converts Dia to Feet
    dia = dia/12
    radius = dia/2
    # Calculate average depth #Need to add Value for Pipe Bedding
    avg_depth = ((start_cov + end_cov) / 2) + dia +.5 

    # Trench width ___________________________________________________________________________________
    # THIS COULD BE CHANGED TO SUIT YOUR TRENCH NEEDS BUT THIS IS GOING OFF WSDOT
    # this is to check underdrain or something of that sort then     applies equation per WSDOT 2.09.4
    if 'underdrain' in desc.lower():
        trench_width_ft = dia + 1
    else: # if pipe ID is equal to or less than 15 inch then This per WSDOT 2.09.4
        if dia <= 1.25 :
            trench_width_ft = dia +(30/12)
        # if pipe ID is equal to or greater than 18 inch then This per WSDOT 2.09.4
        elif dia >= 1.5:
            trench_width_ft = (1.5 * dia)+ 1.5
        else:
            trench_width_ft = 0



    # Trench area_____________________________________________________________________________________
    #THIS COULD BE CHANGED TO SUIT YOUR TRENCH NEEDS BUT THIS IS CALCULATING THE SIDE AREA OF THE TRENCH SO IT WOULD BE A RECTANGLE WITH A TRIANGLE ATTCHED SIDE PROFILE OF THE TRENCH
    trench_area = length * avg_depth

    # Excavation volume (in cubic yards) width x side profile
    excavation_cy = (trench_area * trench_width_ft) / 27

    # Shoring (one face of the trench wall, only if depth > 4')
    if (start_cov > 4) or (end_cov > 4):
        shoring_sf = length * avg_depth 
    else:
        shoring_sf = 0
    # Material Check 
    # List of possible materials to extract from description
    materials = material_keywords

    # Extract material from the description
    found_material = None
    for material in materials:
        if re.search(r'\b' + re.escape(material) + r'\b', desc):  # Check if material exists in desc
            found_material = material
            break

    # If no material found, assign a default or raise an error
    if not found_material:
        Material_check = "Unknown Material"
    else:
        Material_check = "Material Known"
   

    # Pipe cover check logic
    pipe_cover = min(start_cov, end_cov)  # Determine the minimum cover

    if pipe_cover < 2 and found_material == "DI":
        Depth_check = "GOOD"
    elif pipe_cover >= 2 and found_material != "DI":
        Depth_check = "GOOD"
    else:
        Depth_check = "CHECK DEPTH"
    #"DI", "HDPE", "PVC", "CONC", "SD","DIP","PE", "CMP", "SS","PS", "PSS", "SSS", "W", "WM" 
# PIPE ZONE BEDDING AND BACK FILL BASED ON WSDOT STANDARD PLAN B-55.20-03
    #UPDATE FOR OUT DIAMETERS ONCE THAT HAS BEEN IMPLEMENTED

    if found_material  in ("DI", "CONC"):
        CornerAngle = math.asin((radius - (dia*.15))/radius)
        PipeAL = ((180 - math.degrees(CornerAngle)*2)/360)*math.pi*((dia*dia/4)) - (((0.5*dia*math.sin(CornerAngle))*0.35*dia)/2)                                            
        PipeAU =  math.pi*((dia*dia/4)) - PipeAL
        gravel_PZ_bedding = ((trench_width_ft * (0.5+(.15*dia))) -PipeAL)*length/27
        PZ_backfill = ((trench_width_ft * (0.5+(.85*dia))) -PipeAU)*length/27
    if found_material  in ("CMP", "PE"):
        PipeAL = (180/360)*math.pi*((dia*dia/4))                                         
        PipeAU =  math.pi*((dia*dia/4)) - PipeAL
        gravel_PZ_bedding = ((trench_width_ft * (0.5+(.5*dia))) -PipeAL)*length/27
        PZ_backfill = ((trench_width_ft * (0.5+(.5*dia))) -PipeAU )*length/27
    if found_material  in ("HDPE", "PVC","SD"):
        PipeA = math.pi*((dia*dia/4))                                         
        gravel_PZ_bedding = ((trench_width_ft * (0.5+0.5+(dia))) - PipeA)*length/27
        PZ_backfill = 0


    if 'EX' in desc.upper():
        log(f"Skipped existing pipe {name}")
        return ([ name, 'SKIPPED EX PIPE', 0, 0, 0, start_struct, 0, end_struct,0, 0, 0, 0, 0, 0, 0,0,0, 'SKIPPED EX PIPE'])

    else:
        return([
            name, desc, round(dia,2), round(length,1), slope, start_struct, round(start_inv,2), end_struct, round(end_inv,2), round(start_cov,2), round(end_cov,2), round(avg_depth,2), 
            round(trench_width_ft,2), round(excavation_cy,2), round(shoring_sf,2), round(gravel_PZ_bedding),round(PZ_backfill), Material_check, Depth_check
        ])

#-----------------------------------------------------------------------------------------------------------Quanity Calcs Function(s)----------------------------------------------------------------------------
def parse_type_and_diameter(type_str):
    try:
        # Match "Type X" and optionally a diameter like 48, 48", or 48.0
        type_match = re.search(r'[Tt]ype\s*(\d+)', type_str)
        diameter_match = re.search(r'(\d+(?:\.\d+)?)\s*["inIN]*', type_str)

        type_val = type_match.group(1) if type_match else None
        diameter = int(float(diameter_match.group(1))) if diameter_match else None

        return type_val, diameter
    except Exception as e:
        return None, None
        
def QuanityCountStructures(row):
    name = str(row['Name']).lower()
    type_raw = str(row['Type'])
    type_val, diameter = parse_type_and_diameter(type_raw)

    material = None

    if 'ex' in name:
        return

    if 'mh' in name and type_val and diameter:
        material = f"Manhole {diameter} In. Diam. Type {type_val}"
    elif 'cb' in name and type_val:
        if type_val == '2' and diameter:
            material = f"Catch Basin Type 2 {diameter} In. Diam."
        else:
            material = f"Catch Basin Type {type_val}"
    elif 'inl' in name and type_val:
        material = f"Inlet Type {type_val}"
    return material
##############################################################################################################################MAIN_CAD_ESTIMATE#####################################################################################################################################
def CADToEstimate(file_path, log=print, location=None ):
    
    log(f"Running estimation on: {file_path}")
    input_file = file_path
    if not input_file:
        log("No file selected!")
    
    material_keywords = ["DI", "HDPE", "PVC", "CONC", "SD","DIP","PE", "CMP", "SS","PS", "PSS", "SSS", "W", "WM" ]  # Add more as necessary
    PI_df = load_excel_range(input_file, "PipeInput", 1000, 55)
    SI_df = load_excel_range(input_file, "StructureInput", 1000, 23)
    MH_df = load_excel_range(input_file, "MHDataWSDOT", 900, 8)
    CB_df = load_excel_range(input_file, "CBDataWSDOT", 900, 10)
    INL_df = load_excel_range(input_file, "InletDataSPU", 90, 10)
    SI_df['Insertion Rim Elevation'] = SI_df['Insertion Rim Elevation'].apply(cleanInchFeet)
    PI_df['Start Invert Elevation'] =  PI_df['Start Invert Elevation'].apply(cleanInchFeet)
    PI_df['End Invert Elevation'] =  PI_df['End Invert Elevation'].apply(cleanInchFeet)

    calculated_data = []
    for _, row in SI_df.iterrows():
        StructureOuput = StructureCalcs(row,PI_df,CB_df,INL_df,MH_df,log=log)
        calculated_data.append(StructureOuput)
  
    output_columns = [
              'Name', 'Type', 'Max Pipe Depth Ft', 'Area of Structure SF',
              'Structure Depth Ft', 'Extra Area SF', 'Excavation CY', 'Structure Bedding CY'
           ]

           # SORTING DATAFRAME

    def natural_sort_key(s):
        return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(s))]

    StructureOut_df = pd.DataFrame(calculated_data, columns=output_columns)
    final_df_sorted = StructureOut_df.sort_values(by='Name', key=lambda col: col.map(natural_sort_key))

    #======================Writing  the clean StructureInput TO A Excel file ============================#
    with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        final_df_sorted.to_excel(writer, sheet_name="StructureOut", index=False)

#-------------------------------------------Pipes------------------------------------------------------------------
    pipe_results = []
    for col in ['Start Invert Elevation', 'End Invert Elevation', 'Start Cover', 'End Cover', 'Inner Diameter', '2D Length']:
                    PI_df[col] = PI_df[col].apply(cleanInchFeet)
    for _, row in PI_df.iterrows():
        PipeOutput = PipeCalc(row,SI_df,material_keywords,log=log)
        pipe_results.append(PipeOutput) 

    pipe_output_columns = [
                    'Name', 'Description', 'Inner Diameter (FT)', '2D Length (LF)', 'Slope', 'Start Structure',
                    'Start Invert Elevation', 'End Structure', 'End Invert Elevation', 'Start cover', 'End cover',
                    'Avg Depth (FT)', 'Trench Width (FT)', 'Excavation (CY)', 'Shoring (SF)','Gravel Backfill for Pipe Zone Bedding (CY)','Pipe Zone Backfill (CY)', 'Material Check', 'Depth Check'
                ]
    # Display result as DataFrame
    PipeOut_df = pd.DataFrame(pipe_results, columns=pipe_output_columns)
    with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        PipeOut_df.to_excel(writer, sheet_name="PipeOut", index=False) 

#----------------------------------------Qaunity Calculations--------------------------------------------------------------
    structure_material_counts = {}
    for _, row in StructureOut_df.iterrows():
        material = QuanityCountStructures(row)

        if material:
            structure_material_counts[material] = structure_material_counts.get(material, 0) + 1

    #---------------------Pipe Qaunities---------------
    
    PipeOut_df['Inner Diameter (FT)'] = pd.to_numeric(PipeOut_df['Inner Diameter (FT)'], errors='coerce').fillna(0)
    PipeOut_df['2D Length (LF)'] = pd.to_numeric(PipeOut_df['2D Length (LF)'], errors='coerce').fillna(0)

    # Convert feet to inches and group by Description and Diameter
    PipeOut_df['Diameter In'] = PipeOut_df['Inner Diameter (FT)'] * 12
    pipe_groups = PipeOut_df.groupby(['Description', 'Diameter In'])['2D Length (LF)'].sum().reset_index()

    pipe_material_counts = {}
    for _, row in PipeOut_df.iterrows():
        desc = str(row['Description']).strip().upper() 
        diameter = int(round(row['Diameter In']))
        length = round(row['2D Length (LF)'], 2)

        if  'EX' in desc:
            continue
        else:
            # Find material keyword in description
            material_type = None
            for keyword in material_keywords:
                if keyword in desc:
                    material_type = keyword
                    break

            if not material_type:
                material_type = 'Unknown'  # or continue, depending on if you want to skip

            desc = desc.title()  # restore casing for output
            if desc and diameter > 0 and length > 0:
                item = f"{material_type} {diameter} In. Diam."
                pipe_material_counts[item] = pipe_material_counts.get(item, 0) + length

    # --- COMBINE RESULTS ------------------------------------------------------------------
    material_list = []

    for material, qty in structure_material_counts.items():
        material_list.append({
            'Material': material,
            'Quantity': round(qty,1),
            'Unit': 'Each'
        })

    for material, qty in pipe_material_counts.items():
        material_list.append({
            'Material': material,
            'Quantity': round(qty,1),
            'Unit': 'LF'
        })

    material_list_df = pd.DataFrame(material_list)

    # Load workbook and target sheet
    wb = openpyxl.load_workbook(input_file)
    ws = wb['QuantityCost']

    # Starting row and column (adjust as needed)
    start_row = 37  # Excel is 1-based
    start_col = 3   # Column D

    # Write headers manually
    headers = ['Material', 'Quantity', 'Unit']
    for col_offset, header in enumerate(headers):
        ws.cell(row=start_row, column=start_col + col_offset, value=header)

    # Write data rows
    for i, row in material_list_df.iterrows():
        for j, value in enumerate(row):
            ws.cell(row=start_row + i + 1, column=start_col + j, value=value)

    # Save
    wb.save(input_file)

    return 